from fpdf import FPDF
from openpyxl import Workbook
from docx import Document

def txt_to_pdf(filename):
    with open(filename, "r") as file:
        lines = file.readlines()

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    for line in lines:
        pdf.cell(200, 10, txt=line.encode('latin-1', 'replace').decode('latin-1'), ln=True)

    pdf_output = filename.replace(".txt", ".pdf")
    pdf.output(pdf_output)
    print(f"'{pdf_output}' olarak kaydedildi.")

def txt_to_excel(filename):
    with open(filename, "r") as file:
        lines = file.readlines()

    wb = Workbook()
    ws = wb.active

    for index, line in enumerate(lines, 1):
        ws[f"A{index}"] = line.strip()

    excel_output = filename.replace(".txt", ".xlsx")
    wb.save(excel_output)
    print(f"'{excel_output}' olarak kaydedildi.")

def txt_to_word(filename):
    with open(filename, "r") as file:
        lines = file.readlines()

    doc = Document()

    for line in lines:
        doc.add_paragraph(line.strip())

    doc_output = filename.replace(".txt", ".docx")
    doc.save(doc_output)
    print(f"'{doc_output}' olarak kaydedildi.")

if __name__ == "__main__":
    filename = input("Dönüştürmek istediğiniz .txt dosyasının adını girin: ")

    print("\nLütfen dönüşüm formatını seçin:")
    print("1. PDF")
    print("2. XLSX")
    print("3. DOCX")
    choice = input("Seçiminizi yapın (1/2/3): ")

    if choice == "1":
        txt_to_pdf(filename)
    elif choice == "2":
        txt_to_excel(filename)
    elif choice == "3":
        txt_to_word(filename)
    else:
        print("Geçersiz seçim.")
              
